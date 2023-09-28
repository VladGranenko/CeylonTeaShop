import pandas as pd
#
from datetime import datetime
#
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
#
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.colors import Color
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
# models
from models_pack import dBase, ExpenseInvoices, ProductsAndServices, IncomeInvoices, Suppliers


#------------------------------------------ ADMIN-PANEL ---------------------------------------------------#
# ----------------------------------------- LOAD EXCEL  ---------------------------------------------------#
def make_invoice(file_excel):
    model_field = ['name', 'price', 'quantity', 'supplier', 'is_product', 'definition']
    check_list = [False for _ in range(len(model_field))]
    #
    data_excel = pd.read_excel(file_excel)
    column_names = data_excel.columns.tolist()
    #
    if len(model_field) == len(column_names):
        for item in range(len(model_field)):
            if model_field[item] == column_names[item]:
                check_list[item] = True
            else:
                check_list[item] = False
        #
        if all(check_list):
            name = data_excel[column_names[0]]
            price = data_excel[column_names[1]]
            quantity = data_excel[column_names[2]]
            supplier = data_excel[column_names[3]]
            is_product = data_excel[column_names[4]]
            definition = data_excel[column_names[5]]
            return {
                    "name": name, "price": price, "quantity": quantity,
                    "supplier": supplier, "is_product": is_product, 'definition': definition,
                    'length': data_excel.shape[0]
            }
    else:
        return False


def handler_income_inv(file_excel):
    if make_invoice(file_excel):
        invoice = make_invoice(file_excel)
        existing_supplier = dBase.session.query(Suppliers).filter_by(name=invoice["supplier"][0]).first()
        if existing_supplier:
            new_supplier = existing_supplier
        else:
            new_supplier = Suppliers(name=invoice["supplier"][0])
            dBase.session.add(new_supplier)
            dBase.session.commit()
        #
        new_invoice = IncomeInvoices(
                                      total_price=float(invoice["price"].sum()),
                                      total_amount=int(invoice["quantity"].sum()),
                                      supplier_id=new_supplier.id
                                    )
        #
        dBase.session.add(new_invoice)
        dBase.session.commit()
        #
        records_to_insert = []
        for index in range(invoice['length']):
            record = {
                "name": invoice["name"][index],
                "price": float(invoice["price"][index]),
                "quantity": int(invoice["quantity"][index]),
                "supplier": invoice["supplier"][index],
                "is_product": invoice["is_product"][index],
                'definition': invoice['definition'][index],
                "invoice_id": new_invoice.id if invoice["is_product"][index] == True else 0
            }
            records_to_insert.append(record)
        return records_to_insert
    else:
        return False

# ------------------------------------------ GENERATE REPORT  --------------------------------------------#

def check_nodes_range(start_point, finish_point):
    # LIST EXPENSE INVOICES
    list_exp_inv_range = dBase.session.query(ExpenseInvoices).filter(
        ExpenseInvoices.exp_invoice_date >= start_point,
        ExpenseInvoices.exp_invoice_date <= finish_point
    ).all()
    #
    if list_exp_inv_range:
        list_exp_inv_future = dBase.session.query(ExpenseInvoices).filter(
            ExpenseInvoices.exp_invoice_date > finish_point
        ).all()
        # LIST INCOME INVOICES
        list_income_inv = dBase.session.query(IncomeInvoices).filter(
            IncomeInvoices.inc_invoice_date <= finish_point
        ).all()
        #
        dict_column, total_price = dict(), float()
        for invoice in list_exp_inv_range:
            for item in invoice.expense_items:
                if item.name_item not in dict_column.keys():
                    dict_column.setdefault(item.name_item, dict())
                    total_price += float(item.quantity_item * item.price_item)
                    dict_column[item.name_item] = {
                                                    'Кiлькiсть': int(item.quantity_item),
                                                    'Сумма': float(item.quantity_item * item.price_item),
                                                    'Залишок товару': 0
                }
                elif item.name_item in dict_column.keys():
                    total_price += float(item.quantity_item * item.price_item)
                    dict_column[item.name_item]['Кiлькiсть'] += int(item.quantity_item)
                    dict_column[item.name_item]['Сумма'] += float(item.quantity_item * item.price_item)
        #
        for invoice in list_income_inv:
            for item in invoice.products_rel:
                if item.name in dict_column.keys():
                    dict_column[item.name]['Залишок товару'] += item.quantity
        #
        for invoice in list_exp_inv_future:
            for item in invoice.expense_items:
                if item.name_item in dict_column.keys():
                    dict_column[item.name_item]['Залишок товару'] += item.quantity_item
        #
        dict_column[' '] = {
            'Кiлькiсть': ' ',
            'Сумма': total_price,
            'Залишок товару': ' '
        }
        list_from_dict = [
                            dict(zip(['Назва', 'Кiлькiсть', 'Сумма', 'Залишок товару'],
                                     (
                                         key,
                                         dict_column[key]['Кiлькiсть'],
                                         dict_column[key]['Сумма'],
                                         dict_column[key]['Залишок товару']))) for key in dict_column.keys()
        ]
        return list_from_dict
    else:
        return False


def make_report_range(start_date, finish_date, buffer):
    font_size, width = 20, 40
    list_nodes = check_nodes_range(start_point=start_date, finish_point=finish_date)
    if list_nodes:
        report = pd.DataFrame(list_nodes)
        workbook = Workbook()
        ws = workbook.active
        #
        list_col = ['A', 'B', 'C', 'D']
        header = f'Звіт про продажі за період з {start_date} до {finish_date}'
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(list_col))
        cell = ws.cell(row=1, column=1, value=header)
        cell.font = Font(bold=True, size=font_size)
        cell.alignment = Alignment(horizontal='center')
        #
        for row in dataframe_to_rows(report, index=False, header=True):
            ws.append(row)
        #
        fill_row = PatternFill(start_color="B4F5E5", fill_type="solid")
        fill_last_row = PatternFill(start_color="FCAC8D", fill_type="solid")
        #
        for letter in list_col:
            if letter in list_col[1:3]:
                ws.column_dimensions[letter].width = width * 0.5
            elif letter == list_col[0]:
                ws.column_dimensions[letter].width = width * 1.2
            else:
                ws.column_dimensions[letter].width = width
        #
        for row_number, cell in enumerate(ws.iter_rows(min_row=2, min_col=0, max_col=len(list_col)), start=2):
            for cell_in_col in cell:
                cell_in_col.font = Font(bold=True, size=font_size - 4)
                cell_in_col.alignment = Alignment(horizontal='center', vertical='center')
                cell_in_col.fill = fill_row
                if row_number == ws.max_row:
                    cell_in_col.font = Font(bold=True, size=font_size)
                    cell_in_col.fill = fill_last_row
        #
        workbook.save(buffer)
        return workbook
    else:
        return False


#------------------------------------------ SHOP-PANEL ---------------------------------------------------#
# Set Font
font_path = 'fonts/arial.ttf'
pdfmetrics.registerFont(TTFont('ArialUnicode', font_path))


def make_expense_inv(buffer, id_inv):
    pdf_doc = SimpleDocTemplate(buffer, pagesize=letter,  topMargin=inch/4, bottomMargin=inch/2)
    inv_data = ExpenseInvoices.query.filter(ExpenseInvoices.id == id_inv).all()
    styles = getSampleStyleSheet()
    # Title
    title_style = styles["Heading1"]
    title_style.spaceAfter = 40
    title_style.fontSize = 40
    title_style.fontName = 'ArialUnicode'
    title_style.leftIndent = 20
    title = Paragraph("Ceylon Tea Shop", title_style)
    # Number invoice and date
    info_inv = styles['Heading2']
    info_inv.fontName = 'ArialUnicode'
    info_inv.spaceAfter = 2
    info_inv.spaceBefore = 2
    info_inv.leftIndent = 20
    #
    invoice_number_date = Paragraph(
        f"Номер накладної: {7023 + id_inv}<br/>Дата: {datetime.now().strftime('%d.%m.%Y')}", info_inv
    )
    customer = Paragraph(f'Замовник: {inv_data[0].customer.email}', info_inv)
    # title table
    pgph_style = styles['Heading3']
    pgph_style.alignment = 1
    pgph_style.spaceBefore = 20
    pgph_style.spaceAfter = 10
    pgph_style.fontName = 'ArialUnicode'
    #
    title_tbl_product = Paragraph('Лист замовлень', pgph_style)
    title_tbl_services = Paragraph('Лист послуг', pgph_style)

    #------------------------------------ TABLES -----------------------------------------------#
    full_quantity, full_price = 0, 0
    #
    list_service = [item.name for item in
                    ProductsAndServices.query.filter(ProductsAndServices.is_product == False).all()]
    #
    head_products = ['n/n', 'Назва товару', 'Цiна', 'Кількість', "Загальна вартість"]
    head_services = [' ', 'Назва послуги', 'Цiна', ' ', ' ']
    #
    data_table_product, data_table_service = [head_products, ], [head_services, ]
    #
    for ind, value in enumerate(inv_data[0].expense_items):
        row = [
                f'{ind + 1}',
                f'{value.name_item}',
                f'{value.price_item}',
                f'{value.quantity_item}',
                f'{value.quantity_item * value.price_item}'
        ]
        if row:
            if value.name_item in list_service:
                data_table_service.append([' '] + row[1:3] + [' ' * 2])
                full_price += value.price_item

            elif value.name_item not in list_service:
                data_table_product.append(row)
                full_quantity += value.quantity_item
                full_price += value.quantity_item * value.price_item
        else:
            empty = [' ' for _ in range(5)]
            data_table_product.append(empty)
    # create table and add data
    col_widths = [30, 170, 60, 60, 100]
    table_product = Table(data_table_product, colWidths=col_widths)
    table_service = Table(data_table_service, colWidths=col_widths)

    # ------------------------------------ SIGNATURE ------------------------------------#
    signature = [
        ['Замовник', 'Продавець', 'Загальна кількість', 'Загальна вартість'],
        [' ', ' ', str(full_quantity), str(full_price)]
    ]
    table_sgn = Table(signature, colWidths=[110, 110, 100, 100])
    table_sgn.spaceBefore = 20
    # setting
    main_set = [
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, Color(0, 0, 0)),
        ('FONTNAME', (0, 0), (-1, 0), 'ArialUnicode'),
    ]
    # set style for tables
    style = TableStyle(main_set)
    table_product.setStyle(style)
    table_product.keepTogether = 1  # if table couldn't pack in one page (size: 210x297)
    table_service.setStyle(style)
    table_sgn.setStyle(style)
    # Assembler
    elements = [
                 title, invoice_number_date, customer, table_sgn, title_tbl_product, table_product,
                 title_tbl_services, table_service
    ]
    pdf_doc.build(elements)
    return pdf_doc


